import openpyxl

wb = openpyxl.Workbook()
wb.save('work01.xlsx')

wb = openpyxl.load_workbook('work01.xlsx')

s1 = wb['Sheet']
s2=wb.active

print(s1.title, s1.max_row, s1.max_column) 
print(s2.title, s2.max_row, s2.max_column)  

print(s1.sheet_properties)  

s1['A1'].value = 'apple'  
s1['A2'].value = 'orange' 
s1['A3'].value = 'banana' 
s1.cell(1,2).value = 100 
s1.cell(2,2).value = 200
s1.cell(3,2).value = 300

wb.save('work01.xlsx')

data = [[1,2,3],[4,5,6],[7,8,9]] 
for i in data:
    s1.append(i)

wb.save('work01.xlsx')

wb.create_sheet("Sheet 2")
s2 = wb['Sheet 2']

data = [[1,2],[3,4]]  
for y in range(len(data)):
    for x in range(len(data[y])):
        row = 2 + y 
        col = 2 + x  
        s2.cell(row, col).value = data[y][x]

wb.save('work01.xlsx')

from openpyxl.styles import Font, PatternFill 

s1['A7'] = '=sum(a4:a6)'  
s1['A6'].font = Font(name='Arial', color='ff0000', size=30, bold=True)

c=s1['C4']
s1['D4'].value=c.value

wb.save('work01.xlsx')


